import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'c:\Users\Jahidul Islam\Downloads\jahidul\Employee Performance Evalution Sheet.xlsx',
    data_only=True
)
ws = wb['Kam Data']

geo_members = ['Jahidul', 'Sabit', 'Komal', 'Hasibul', 'Shourav', 'Roni']

stats = {}
for m in geo_members:
    stats[m] = {
        'WIP': 0, 'Revision': 0, 'Delivered': 0,
        'Cancelled': 0, 'Total': 0,
        'DeliveredAmt': 0.0, 'WIPAmt': 0.0, 'AmtErrors': 0
    }

for row in ws.iter_rows(min_row=2, max_row=1500, values_only=True):
    assign = str(row[18]).strip() if row[18] else ''   # Col S
    status = str(row[19]).strip() if row[19] else ''   # Col T
    amt_x = row[23]   # Col X — Delivered & WIP দুটোর জন্যই X ব্যবহার

    for m in geo_members:
        if m.lower() in assign.lower():
            stats[m]['Total'] += 1
            slash_count = max(assign.count('/'), 0) + 1

            if status == 'WIP' or status == 'Revision':
                if status == 'WIP':
                    stats[m]['WIP'] += 1
                else:
                    stats[m]['Revision'] += 1
                try:
                    if amt_x not in (None, '', 'None'):
                        val = float(str(amt_x).replace('$','').replace(',','').strip())
                        stats[m]['WIPAmt'] += val / slash_count
                except:
                    stats[m]['AmtErrors'] += 1

            elif status == 'Delivered':
                stats[m]['Delivered'] += 1
                try:
                    if amt_x not in (None, '', 'None'):
                        # ✅ X column থেকে Delivered amount নিচ্ছি
                        val = float(str(amt_x).replace('$','').replace(',','').strip())
                        stats[m]['DeliveredAmt'] += val / slash_count
                except:
                    stats[m]['AmtErrors'] += 1

            elif status == 'Cancelled':
                stats[m]['Cancelled'] += 1

print("=" * 105)
print("  GEO RANKERS — Data Check  [Delivered $ = Col X | WIP $ = Col X]")
print("=" * 105)
print()
print(f"{'Member':<12} | {'Total':>5} | {'WIP':>4} | {'Revision':>8} | {'Delivered':>9} | {'Cancelled':>9} | {'Delivered $ (X)':>15} | {'WIP Pipeline $ (X)':>18}")
print("-" * 105)
for m in geo_members:
    s = stats[m]
    print(f"{m:<12} | {s['Total']:>5} | {s['WIP']:>4} | {s['Revision']:>8} | {s['Delivered']:>9} | {s['Cancelled']:>9} | {s['DeliveredAmt']:>15.2f} | {s['WIPAmt']:>18.2f}")

print()
print("=" * 105)
print("  Target ($1,100) vs Delivered $ (Col X)")
print("=" * 105)
print()
print(f"{'Member':<12} | {'Target':>8} | {'Delivered $ (X)':>15} | {'Remaining':>12} | {'WIP Pipeline $':>14} | {'Assessment'}")
print("-" * 85)
for m in geo_members:
    s = stats[m]
    target = 1100.0
    remaining = s['DeliveredAmt'] - target
    if s['DeliveredAmt'] >= target:
        assess = "✅ Target Met!"
    elif s['DeliveredAmt'] >= target * 0.75:
        assess = "⚠️  75%+ Done"
    elif s['DeliveredAmt'] >= target * 0.5:
        assess = "🔶 50%+ Done"
    elif s['DeliveredAmt'] > 0:
        assess = "🔴 Below 50%"
    else:
        assess = "❌ No Delivery"
    print(f"{m:<12} | {target:>8.0f} | {s['DeliveredAmt']:>15.2f} | {remaining:>+12.2f} | {s['WIPAmt']:>14.2f} | {assess}")

print()
print("=== Jahidul-এর Delivered Orders — X column Amount ===")
print()
for row in ws.iter_rows(min_row=2, max_row=1500, values_only=True):
    assign = str(row[18]).strip() if row[18] else ''
    status = str(row[19]).strip() if row[19] else ''
    if 'jahidul' in assign.lower() and status == 'Delivered':
        slash_count = max(assign.count('/'), 0) + 1
        amt_x = row[23]
        try:
            val = float(str(amt_x).replace('$','').replace(',','').strip())
            share = val / slash_count
        except:
            val = 'N/A'
            share = 'N/A'
        print(f"  Order : {row[13]}")
        print(f"  Client: {row[10]}")
        print(f"  Assign: {assign}  (÷{slash_count})")
        print(f"  Col X Amount: ${val}  →  Jahidul Share: ${share}")
        print("-" * 55)
